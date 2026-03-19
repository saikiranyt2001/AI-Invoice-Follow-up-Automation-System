from __future__ import annotations

import csv
from datetime import date
from io import BytesIO, StringIO

from openpyxl import load_workbook

REQUIRED_FIELDS = {"customer_name", "customer_email", "amount", "due_date"}
FIELD_ALIASES = {
    "customer": "customer_name",
    "name": "customer_name",
    "email": "customer_email",
    "phone": "customer_phone",
    "mobile": "customer_phone",
    "invoice_amount": "amount",
    "total": "amount",
    "due": "due_date",
    "due date": "due_date",
}


def _normalize_key(value: object) -> str:
    return str(value or "").strip().lower().replace("-", "_").replace(" ", "_")


def _canonical_key(value: object) -> str:
    normalized = _normalize_key(value)
    return FIELD_ALIASES.get(normalized, normalized)


def _normalize_row(raw_row: dict[object, object]) -> dict[str, str]:
    row: dict[str, str] = {}
    for key, value in raw_row.items():
        row[_canonical_key(key)] = str(value or "").strip()
    return row


def parse_invoice_file(filename: str, content: bytes) -> list[dict[str, str]]:
    lowered = filename.lower()
    if lowered.endswith(".csv"):
        return _parse_csv(content)
    if lowered.endswith(".xlsx"):
        return _parse_excel(content)
    if lowered.endswith(".xls"):
        raise ValueError("Legacy .xls files are not supported. Save as .xlsx and upload again.")
    raise ValueError("Unsupported file type. Upload a CSV or Excel (.xlsx) file.")


def validate_invoice_rows(rows: list[dict[str, str]]) -> list[str]:
    errors: list[str] = []
    for index, row in enumerate(rows, start=2):
        missing = sorted(field for field in REQUIRED_FIELDS if not row.get(field))
        if missing:
            errors.append(f"Row {index}: missing {', '.join(missing)}")
            continue

        try:
            amount = float(row["amount"])
            if amount <= 0:
                raise ValueError()
        except Exception:
            errors.append(f"Row {index}: amount must be a number greater than 0")

        try:
            date.fromisoformat(row["due_date"])
        except Exception:
            errors.append(f"Row {index}: due_date must be in YYYY-MM-DD format")
    return errors


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV file must be UTF-8 encoded") from exc

    reader = csv.DictReader(StringIO(text))
    headers = {_canonical_key(name) for name in (reader.fieldnames or [])}
    if not REQUIRED_FIELDS.issubset(headers):
        raise ValueError("File must include customer_name, customer_email, amount, due_date")
    return [
        _normalize_row(row)
        for row in reader
        if any(str(value or "").strip() for value in row.values())
    ]


def _parse_excel(content: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_canonical_key(cell) for cell in rows[0]]
    if not REQUIRED_FIELDS.issubset(set(headers)):
        raise ValueError("File must include customer_name, customer_email, amount, due_date")

    parsed_rows: list[dict[str, str]] = []
    for values in rows[1:]:
        if not values or not any(value not in (None, "") for value in values):
            continue
        raw = {
            headers[index]: values[index] if index < len(values) else ""
            for index in range(len(headers))
        }
        parsed_rows.append(_normalize_row(raw))
    return parsed_rows
